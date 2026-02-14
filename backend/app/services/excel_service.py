"""Excel generation service for Malar Market Digital Ledger."""

import io
import logging
from datetime import datetime, date
from typing import Dict, List, Optional, Any
from decimal import Decimal

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    Protection, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

from app.services.report_templates import get_template, ExcelStyle

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for generating Excel reports with professional formatting."""
    
    def __init__(self, language: str = "en"):
        """
        Initialize Excel service.
        
        Args:
            language: Language code ('en' or 'ta')
        """
        self.language = language
        self.template = get_template(language)
        self.style = self.template.get_excel_style()
        self.text = self.template.get_text()
    
    def _create_workbook(self) -> openpyxl.Workbook:
        """
        Create a new Excel workbook.
        
        Returns:
            openpyxl Workbook instance
        """
        workbook = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])
        return workbook
    
    def _apply_header_style(
        self,
        cell,
        fill_color: str = None
    ):
        """
        Apply header style to a cell.
        
        Args:
            cell: Excel cell to style
            fill_color: Background fill color
        """
        if fill_color is None:
            fill_color = self.style.header_fill
        
        cell.font = Font(
            name='Calibri',
            size=self.style.header_font_size,
            bold=self.style.header_font_bold,
            color=self.style.header_font_color
        )
        cell.fill = PatternFill(
            start_color=fill_color,
            end_color=fill_color,
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        cell.border = Border(
            left=Side(style='thin', color=self.style.border_color),
            right=Side(style='thin', color=self.style.border_color),
            top=Side(style='thin', color=self.style.border_color),
            bottom=Side(style='thin', color=self.style.border_color)
        )
    
    def _apply_body_style(
        self,
        cell,
        is_alternate: bool = False,
        is_number: bool = False,
        is_currency: bool = False
    ):
        """
        Apply body style to a cell.
        
        Args:
            cell: Excel cell to style
            is_alternate: Whether to use alternate row color
            is_number: Whether cell contains a number
            is_currency: Whether cell contains currency
        """
        fill_color = self.style.alternate_fill if is_alternate else None
        
        cell.font = Font(
            name='Calibri',
            size=self.style.body_font_size,
            color=self.style.body_font_color
        )
        
        if fill_color:
            cell.fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type='solid'
            )
        
        alignment = 'right' if (is_number or is_currency) else 'left'
        cell.alignment = Alignment(
            horizontal=alignment,
            vertical='center',
            wrap_text=True
        )
        
        cell.border = Border(
            left=Side(style='thin', color=self.style.border_color),
            right=Side(style='thin', color=self.style.border_color),
            top=Side(style='thin', color=self.style.border_color),
            bottom=Side(style='thin', color=self.style.border_color)
        )
        
        if is_currency:
            cell.number_format = self.style.currency_format
        elif is_number:
            cell.number_format = self.style.number_format
    
    def _apply_date_style(self, cell):
        """
        Apply date style to a cell.
        
        Args:
            cell: Excel cell to style
        """
        cell.font = Font(
            name='Calibri',
            size=self.style.body_font_size,
            color=self.style.body_font_color
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        cell.border = Border(
            left=Side(style='thin', color=self.style.border_color),
            right=Side(style='thin', color=self.style.border_color),
            top=Side(style='thin', color=self.style.border_color),
            bottom=Side(style='thin', color=self.style.border_color)
        )
        cell.number_format = self.style.date_format
    
    def _auto_fit_columns(self, sheet):
        """
        Auto-fit column widths based on content.
        
        Args:
            sheet: Excel worksheet
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add some padding
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _freeze_header(self, sheet):
        """
        Freeze the header row.
        
        Args:
            sheet: Excel worksheet
        """
        sheet.freeze_panes = "A2"
    
    def _add_title_row(
        self,
        sheet,
        title: str,
        merge_range: str = "A1:E1"
    ):
        """
        Add title row to sheet.
        
        Args:
            sheet: Excel worksheet
            title: Title text
            merge_range: Cell range to merge for title
        """
        cell = sheet[merge_range.split(":")[0]]
        cell.value = title
        cell.font = Font(
            name='Calibri',
            size=14,
            bold=True,
            color='FFFFFF'
        )
        cell.fill = PatternFill(
            start_color='374151',
            end_color='374151',
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )
        sheet.merge_cells(merge_range)
    
    def _create_summary_sheet(
        self,
        workbook: openpyxl.Workbook,
        summary_data: Dict[str, Any]
    ):
        """
        Create summary sheet.
        
        Args:
            workbook: Excel workbook
            summary_data: Summary information
        """
        sheet = workbook.create_sheet("Summary", 0)
        
        # Add title
        self._add_title_row(sheet, "Monthly Summary", "A1:B1")
        
        # Add summary data
        data = [
            ["Report Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Farmers", summary_data.get('total_farmers', 0)],
            ["Total Entries", summary_data.get('total_entries', 0)],
            ["Total Weight (kg)", summary_data.get('total_weight', 0)],
            ["Total Gross Value (₹)", summary_data.get('total_gross', 0)],
            ["Total Commission (₹)", summary_data.get('total_commission', 0)],
            ["Total Net Amount (₹)", summary_data.get('total_net', 0)]
        ]
        
        for idx, (label, value) in enumerate(data, start=2):
            label_cell = sheet[f"A{idx}"]
            value_cell = sheet[f"B{idx}"]
            
            label_cell.value = label
            value_cell.value = value
            
            self._apply_body_style(label_cell)
            self._apply_body_style(value_cell, is_number=isinstance(value, (int, float, Decimal)))
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)
        
        # Set row heights
        sheet.row_dimensions[1].height = self.style.header_row_height
    
    def _create_farmer_summaries_sheet(
        self,
        workbook: openpyxl.Workbook,
        farmer_summaries: List[Dict[str, Any]]
    ):
        """
        Create farmer summaries sheet.
        
        Args:
            workbook: Excel workbook
            farmer_summaries: List of farmer summary data
        """
        sheet = workbook.create_sheet("Farmers", 1)
        
        # Add title
        self._add_title_row(sheet, "Farmer Summaries", "A1:J1")
        
        # Add header row
        headers = [
            "Farmer ID",
            "Farmer Name",
            "Village",
            "Phone",
            "Total Entries",
            "Total Weight (kg)",
            "Gross Value (₹)",
            "Commission (₹)",
            "Advances (₹)",
            "Net Amount (₹)"
        ]
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=idx)
            cell.value = header
            self._apply_header_style(cell)
        
        # Add data rows
        for row_idx, farmer in enumerate(farmer_summaries, start=3):
            data = [
                farmer.get('farmer_code', ''),
                farmer.get('name', ''),
                farmer.get('village', ''),
                farmer.get('phone', ''),
                farmer.get('total_entries', 0),
                farmer.get('total_quantity', 0),
                farmer.get('gross_amount', 0),
                farmer.get('total_commission', 0),
                farmer.get('total_advances', 0),
                farmer.get('net_payable', 0)
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                is_alternate = (row_idx % 2 == 0)
                is_number = col_idx >= 5
                is_currency = col_idx in [7, 8, 9, 10]
                
                self._apply_body_style(
                    cell,
                    is_alternate=is_alternate,
                    is_number=is_number,
                    is_currency=is_currency
                )
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)
        
        # Freeze header
        self._freeze_header(sheet)
        
        # Set row heights
        sheet.row_dimensions[1].height = self.style.header_row_height
        sheet.row_dimensions[2].height = self.style.header_row_height
    
    def _create_daily_entries_sheet(
        self,
        workbook: openpyxl.Workbook,
        entries: List[Dict[str, Any]]
    ):
        """
        Create daily entries sheet.
        
        Args:
            workbook: Excel workbook
            entries: List of daily entry data
        """
        sheet = workbook.create_sheet("Daily Entries", 2)
        
        # Add title
        self._add_title_row(sheet, "Daily Entries", "A1:I1")
        
        # Add header row
        headers = [
            "Date",
            "Time",
            "Farmer ID",
            "Farmer Name",
            "Flower Type",
            "Quantity (kg)",
            "Rate (₹)",
            "Total (₹)",
            "Net (₹)"
        ]
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=idx)
            cell.value = header
            self._apply_header_style(cell)
        
        # Add data rows
        for row_idx, entry in enumerate(entries, start=3):
            data = [
                entry.get('entry_date', ''),
                entry.get('entry_time', ''),
                entry.get('farmer_code', ''),
                entry.get('farmer_name', ''),
                entry.get('flower_name', ''),
                entry.get('quantity', 0),
                entry.get('rate_per_unit', 0),
                entry.get('total_amount', 0),
                entry.get('net_amount', 0)
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                is_alternate = (row_idx % 2 == 0)
                is_number = col_idx >= 6
                is_currency = col_idx in [7, 8, 9]
                is_date = col_idx == 1
                
                if is_date:
                    self._apply_date_style(cell)
                else:
                    self._apply_body_style(
                        cell,
                        is_alternate=is_alternate,
                        is_number=is_number,
                        is_currency=is_currency
                    )
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)
        
        # Freeze header
        self._freeze_header(sheet)
        
        # Set row heights
        sheet.row_dimensions[1].height = self.style.header_row_height
        sheet.row_dimensions[2].height = self.style.header_row_height
    
    def _create_cash_advances_sheet(
        self,
        workbook: openpyxl.Workbook,
        advances: List[Dict[str, Any]]
    ):
        """
        Create cash advances sheet.
        
        Args:
            workbook: Excel workbook
            advances: List of cash advance data
        """
        sheet = workbook.create_sheet("Cash Advances", 3)
        
        # Add title
        self._add_title_row(sheet, "Cash Advances", "A1:G1")
        
        # Add header row
        headers = [
            "Date",
            "Farmer ID",
            "Farmer Name",
            "Amount (₹)",
            "Reason",
            "Status",
            "Approved By"
        ]
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=idx)
            cell.value = header
            self._apply_header_style(cell)
        
        # Add data rows
        for row_idx, advance in enumerate(advances, start=3):
            status = self.template.get_status_text(advance.get('status', 'pending'))
            
            data = [
                advance.get('advance_date', ''),
                advance.get('farmer_code', ''),
                advance.get('farmer_name', ''),
                advance.get('amount', 0),
                advance.get('reason', ''),
                status,
                advance.get('approved_by', '')
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                is_alternate = (row_idx % 2 == 0)
                is_currency = col_idx == 4
                is_date = col_idx == 1
                
                if is_date:
                    self._apply_date_style(cell)
                else:
                    self._apply_body_style(
                        cell,
                        is_alternate=is_alternate,
                        is_currency=is_currency
                    )
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)
        
        # Freeze header
        self._freeze_header(sheet)
        
        # Set row heights
        sheet.row_dimensions[1].height = self.style.header_row_height
        sheet.row_dimensions[2].height = self.style.header_row_height
    
    def _create_settlements_sheet(
        self,
        workbook: openpyxl.Workbook,
        settlements: List[Dict[str, Any]]
    ):
        """
        Create settlements sheet.
        
        Args:
            workbook: Excel workbook
            settlements: List of settlement data
        """
        sheet = workbook.create_sheet("Settlements", 4)
        
        # Add title
        self._add_title_row(sheet, "Settlements", "A1:L1")
        
        # Add header row
        headers = [
            "Settlement No",
            "Date",
            "Farmer ID",
            "Farmer Name",
            "Period Start",
            "Period End",
            "Entries",
            "Weight (kg)",
            "Gross (₹)",
            "Commission (₹)",
            "Net (₹)",
            "Status"
        ]
        
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=idx)
            cell.value = header
            self._apply_header_style(cell)
        
        # Add data rows
        for row_idx, settlement in enumerate(settlements, start=3):
            status = self.template.get_status_text(settlement.get('status', 'pending'))
            
            data = [
                settlement.get('settlement_number', ''),
                settlement.get('settlement_date', ''),
                settlement.get('farmer_code', ''),
                settlement.get('farmer_name', ''),
                settlement.get('period_start', ''),
                settlement.get('period_end', ''),
                settlement.get('total_entries', 0),
                settlement.get('total_quantity', 0),
                settlement.get('gross_amount', 0),
                settlement.get('total_commission', 0),
                settlement.get('net_payable', 0),
                status
            ]
            
            for col_idx, value in enumerate(data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                is_alternate = (row_idx % 2 == 0)
                is_currency = col_idx in [9, 10, 11]
                is_date = col_idx in [2, 5, 6]
                
                if is_date:
                    self._apply_date_style(cell)
                else:
                    self._apply_body_style(
                        cell,
                        is_alternate=is_alternate,
                        is_currency=is_currency
                    )
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)
        
        # Freeze header
        self._freeze_header(sheet)
        
        # Set row heights
        sheet.row_dimensions[1].height = self.style.header_row_height
        sheet.row_dimensions[2].height = self.style.header_row_height
    
    def generate_monthly_report(
        self,
        summary_data: Dict[str, Any],
        farmer_summaries: List[Dict[str, Any]],
        daily_entries: List[Dict[str, Any]],
        cash_advances: List[Dict[str, Any]],
        settlements: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate master monthly report Excel file.
        
        Args:
            summary_data: Summary statistics
            farmer_summaries: List of farmer summaries
            daily_entries: List of daily entries
            cash_advances: List of cash advances
            settlements: List of settlements
            
        Returns:
            Excel file bytes
        """
        workbook = self._create_workbook()
        
        # Create sheets
        self._create_summary_sheet(workbook, summary_data)
        self._create_farmer_summaries_sheet(workbook, farmer_summaries)
        self._create_daily_entries_sheet(workbook, daily_entries)
        self._create_cash_advances_sheet(workbook, cash_advances)
        self._create_settlements_sheet(workbook, settlements)
        
        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info("Generated monthly report Excel file")
        
        return excel_bytes
    
    def generate_farmer_statement(
        self,
        farmer_data: Dict[str, Any],
        summary_data: Dict[str, Any],
        entries: List[Dict[str, Any]],
        advances: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate individual farmer statement Excel file.
        
        Args:
            farmer_data: Farmer information
            summary_data: Summary statistics
            entries: List of daily entries
            advances: List of cash advances
            
        Returns:
            Excel file bytes
        """
        workbook = self._create_workbook()
        
        # Add title sheet with farmer details
        sheet = workbook.create_sheet("Farmer Details", 0)
        
        # Add title
        self._add_title_row(sheet, f"Farmer Statement - {farmer_data.get('name', '')}", "A1:D1")
        
        # Add farmer details
        details = [
            ["Farmer ID", farmer_data.get('farmer_code', '')],
            ["Name", farmer_data.get('name', '')],
            ["Village", farmer_data.get('village', '')],
            ["Phone", farmer_data.get('phone', '')],
            ["", ""],
            ["Summary", ""],
            ["Total Weight (kg)", summary_data.get('total_quantity', 0)],
            ["Gross Value (₹)", summary_data.get('gross_amount', 0)],
            ["Commission (₹)", summary_data.get('total_commission', 0)],
            ["Advances (₹)", summary_data.get('total_advances', 0)],
            ["Net Amount (₹)", summary_data.get('net_payable', 0)]
        ]
        
        for idx, (label, value) in enumerate(details, start=2):
            label_cell = sheet[f"A{idx}"]
            value_cell = sheet[f"B{idx}"]
            
            label_cell.value = label
            value_cell.value = value
            
            if label == "":
                continue
            
            if label == "Summary":
                label_cell.font = Font(name='Calibri', size=12, bold=True)
                label_cell.fill = PatternFill(
                    start_color='374151',
                    end_color='374151',
                    fill_type='solid'
                )
                label_cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
                sheet.merge_cells(f"A{idx}:B{idx}")
            else:
                self._apply_body_style(label_cell)
                is_number = isinstance(value, (int, float, Decimal))
                is_currency = label in ["Gross Value (₹)", "Commission (₹)", "Advances (₹)", "Net Amount (₹)"]
                self._apply_body_style(value_cell, is_number=is_number, is_currency=is_currency)
        
        self._auto_fit_columns(sheet)
        
        # Create entries sheet
        if entries:
            entries_sheet = workbook.create_sheet("Daily Entries", 1)
            self._add_title_row(entries_sheet, "Daily Entries", "A1:F1")
            
            headers = ["Date", "Flower Type", "Quantity (kg)", "Rate (₹)", "Total (₹)", "Net (₹)"]
            for idx, header in enumerate(headers, start=1):
                cell = entries_sheet.cell(row=2, column=idx)
                cell.value = header
                self._apply_header_style(cell)
            
            for row_idx, entry in enumerate(entries, start=3):
                data = [
                    entry.get('entry_date', ''),
                    entry.get('flower_name', ''),
                    entry.get('quantity', 0),
                    entry.get('rate_per_unit', 0),
                    entry.get('total_amount', 0),
                    entry.get('net_amount', 0)
                ]
                
                for col_idx, value in enumerate(data, start=1):
                    cell = entries_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    is_alternate = (row_idx % 2 == 0)
                    is_number = col_idx >= 3
                    is_currency = col_idx in [4, 5, 6]
                    is_date = col_idx == 1
                    
                    if is_date:
                        self._apply_date_style(cell)
                    else:
                        self._apply_body_style(
                            cell,
                            is_alternate=is_alternate,
                            is_number=is_number,
                            is_currency=is_currency
                        )
            
            self._auto_fit_columns(entries_sheet)
            self._freeze_header(entries_sheet)
            entries_sheet.row_dimensions[1].height = self.style.header_row_height
            entries_sheet.row_dimensions[2].height = self.style.header_row_height
        
        # Create advances sheet
        if advances:
            advances_sheet = workbook.create_sheet("Cash Advances", 2)
            self._add_title_row(advances_sheet, "Cash Advances", "A1:E1")
            
            headers = ["Date", "Amount (₹)", "Reason", "Status", "Approved By"]
            for idx, header in enumerate(headers, start=1):
                cell = advances_sheet.cell(row=2, column=idx)
                cell.value = header
                self._apply_header_style(cell)
            
            for row_idx, advance in enumerate(advances, start=3):
                status = self.template.get_status_text(advance.get('status', 'pending'))
                
                data = [
                    advance.get('advance_date', ''),
                    advance.get('amount', 0),
                    advance.get('reason', ''),
                    status,
                    advance.get('approved_by', '')
                ]
                
                for col_idx, value in enumerate(data, start=1):
                    cell = advances_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    is_alternate = (row_idx % 2 == 0)
                    is_currency = col_idx == 2
                    is_date = col_idx == 1
                    
                    if is_date:
                        self._apply_date_style(cell)
                    else:
                        self._apply_body_style(
                            cell,
                            is_alternate=is_alternate,
                            is_currency=is_currency
                        )
            
            self._auto_fit_columns(advances_sheet)
            self._freeze_header(advances_sheet)
            advances_sheet.row_dimensions[1].height = self.style.header_row_height
            advances_sheet.row_dimensions[2].height = self.style.header_row_height
        
        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"Generated farmer statement Excel for farmer {farmer_data.get('id')}")
        
        return excel_bytes
    
    def generate_daily_summary(
        self,
        date: date,
        summary_data: Dict[str, Any],
        entries: List[Dict[str, Any]]
    ) -> bytes:
        """
        Generate daily summary Excel file.
        
        Args:
            date: Report date
            summary_data: Summary statistics
            entries: List of daily entries
            
        Returns:
            Excel file bytes
        """
        workbook = self._create_workbook()
        
        # Create summary sheet
        sheet = workbook.create_sheet("Summary", 0)
        
        # Add title
        date_str = date.strftime("%Y-%m-%d")
        self._add_title_row(sheet, f"Daily Summary - {date_str}", "A1:B1")
        
        # Add summary data
        data = [
            ["Date", date_str],
            ["Total Entries", summary_data.get('total_entries', 0)],
            ["Total Weight (kg)", summary_data.get('total_weight', 0)],
            ["Gross Value (₹)", summary_data.get('gross_amount', 0)],
            ["Total Commission (₹)", summary_data.get('total_commission', 0)],
            ["Net Amount (₹)", summary_data.get('net_amount', 0)]
        ]
        
        for idx, (label, value) in enumerate(data, start=2):
            label_cell = sheet[f"A{idx}"]
            value_cell = sheet[f"B{idx}"]
            
            label_cell.value = label
            value_cell.value = value
            
            self._apply_body_style(label_cell)
            is_number = isinstance(value, (int, float, Decimal))
            is_currency = label in ["Gross Value (₹)", "Total Commission (₹)", "Net Amount (₹)"]
            self._apply_body_style(value_cell, is_number=is_number, is_currency=is_currency)
        
        self._auto_fit_columns(sheet)
        
        # Create entries sheet
        if entries:
            entries_sheet = workbook.create_sheet("Entries", 1)
            self._add_title_row(entries_sheet, "Daily Entries", "A1:I1")
            
            headers = [
                "Time",
                "Farmer ID",
                "Farmer Name",
                "Flower Type",
                "Quantity (kg)",
                "Rate (₹)",
                "Total (₹)",
                "Commission (₹)",
                "Net (₹)"
            ]
            
            for idx, header in enumerate(headers, start=1):
                cell = entries_sheet.cell(row=2, column=idx)
                cell.value = header
                self._apply_header_style(cell)
            
            for row_idx, entry in enumerate(entries, start=3):
                data = [
                    entry.get('entry_time', ''),
                    entry.get('farmer_code', ''),
                    entry.get('farmer_name', ''),
                    entry.get('flower_name', ''),
                    entry.get('quantity', 0),
                    entry.get('rate_per_unit', 0),
                    entry.get('total_amount', 0),
                    entry.get('commission_amount', 0),
                    entry.get('net_amount', 0)
                ]
                
                for col_idx, value in enumerate(data, start=1):
                    cell = entries_sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    is_alternate = (row_idx % 2 == 0)
                    is_number = col_idx >= 5
                    is_currency = col_idx in [6, 7, 8, 9]
                    
                    self._apply_body_style(
                        cell,
                        is_alternate=is_alternate,
                        is_number=is_number,
                        is_currency=is_currency
                    )
            
            self._auto_fit_columns(entries_sheet)
            self._freeze_header(entries_sheet)
            entries_sheet.row_dimensions[1].height = self.style.header_row_height
            entries_sheet.row_dimensions[2].height = self.style.header_row_height
        
        # Save to bytes
        buffer = io.BytesIO()
        workbook.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"Generated daily summary Excel for {date_str}")
        
        return excel_bytes
